"""
Скрипт знебарвлює всі комірки Excel-таблиці,
крім тих, що помічені збереженими кольорами.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from config import EXCEL_INDEXED_COLORS, KEEP_COLORS, logger


class ColorFilter:
    """Визначає, чи потрібно зберегти наявну заливку комірки."""

    def __init__(self):
        self.keep_colors = {c.upper() for c in KEEP_COLORS}
        self.indexed_palette = [c.upper() for c in EXCEL_INDEXED_COLORS]

    def current_hex(self, cell) -> str | None:
        """Повертає 6-символьний hex поточної заливки або None."""
        fill = cell.fill
        if fill is None or fill.fill_type in (None, "none"):
            return None

        color = fill.fgColor
        if color is None:
            return None

        if color.type == "rgb":
            rgb = color.rgb.upper()
            return rgb[2:] if len(rgb) == 8 else rgb

        if color.type == "indexed":
            idx = color.indexed
            if isinstance(idx, int) and idx < len(self.indexed_palette):
                return self.indexed_palette[idx]

        return None

    def should_keep(self, cell) -> bool:
        """True — якщо поточний колір комірки є збереженим."""
        hex_color = self.current_hex(cell)
        if hex_color is None:
            return False
        return hex_color in self.keep_colors


class TextColorRule:
    """Одне правило: набір ключових слів → цільовий колір заливки."""

    def __init__(self, target_hex: str, keywords: list[str]):
        self.fill = PatternFill("solid", fgColor=target_hex)
        self.target_hex = target_hex.upper()
        # зберігаємо ключові слова у нижньому регістрі для порівняння
        self.keywords = [kw.lower() for kw in keywords]

    def matches(self, cell) -> bool:
        """True — якщо текст комірки містить хоча б одне ключове слово."""

        value = cell.value  # значення в комірці
        if value is None:
            return False

        text = str(value).lower()
        return any(kw in text for kw in self.keywords)


class ExcelDecolorizer:
    """Завантажує Excel, знебарвлює та зафарбовує комірки за правилами."""

    NO_FILL = PatternFill(fill_type=None)

    def __init__(
        self,
        input_path: str,
        output_path: str,
        color_filter: ColorFilter,
        text_rules: list[TextColorRule],
    ):
        self.input_path = input_path
        self.output_path = output_path
        self.color_filter = color_filter
        self.text_rules = text_rules

        self.total_cleared = 0
        self.total_kept = 0
        self.total_colored_by_rule = 0

    def process(self) -> None:
        """"""
        wb = load_workbook(self.input_path)  # вихідна таблиця для знебарвлення
        for sheet in wb.worksheets:
            self._process_sheet(sheet)
        wb.save(self.output_path)
        self.print_report()

    def _process_sheet(self, sheet) -> None:
        """ """
        for row in sheet.iter_rows():
            for cell in row:
                self._process_cell(cell)

    @staticmethod
    def _is_gradient(cell) -> bool:
        """Перевіряє комірку чи зафарбована градієнтом"""
        fill = cell.fill
        return (
            fill is not None
            and getattr(fill, "fgColor", None) is None
            and getattr(fill, "fill_type", None) not in (None, "none")
        )

    def _process_cell(self, cell) -> None:
        """ """

        if self._is_gradient(cell):
            self.total_kept += 1
            return

        # 1. Текст збігається з правилом → перефарбувати незалежно від поточного кольору
        for rule in self.text_rules:
            if rule.matches(cell):
                cell.fill = rule.fill
                self.total_colored_by_rule += 1
                return

        # 2. Текст не збігається, але колір збережений (жовтий/синій) → залишити
        if self.color_filter.should_keep(cell):
            self.total_kept += 1
            return

        # 3. Все інше → знебарвити
        cell.fill = self.NO_FILL
        self.total_cleared += 1

    def print_report(self) -> str:
        logger.info("Готово! Збережено: %s", self.output_path)
        logger.info("Знебарвлено комірок: %s", self.total_cleared)
        logger.info("Збережено кольорових: %s", self.total_kept)
        logger.info("Зафарбовано за текстом: %s", self.total_colored_by_rule)
        return (
            f"Знебарвлено комірок: {self.total_cleared}\n"
            f"Збережено кольорових: {self.total_kept}\n"
            f"Зафарбовано за текстом: {self.total_colored_by_rule}"
        )
